#!/usr/bin/env python3
#
# Copyright (c) 2019, Arista Networks EOS+
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#
# * Redistributions of source code must retain the above copyright notice, this
#   list of conditions and the following disclaimer.
#
# * Redistributions in binary form must reproduce the above copyright notice,
#   this list of conditions and the following disclaimer in the documentation
#   and/or other materials provided with the distribution.
#
# * Neither the name of the Arista nor the names of its
#   contributors may be used to endorse or promote products derived from
#   this software without specific prior written permission.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
# FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
# DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
# SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
# CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
# OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
# OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
#

"""Utilities for using PyTest in network testing"""

import logging
import yaml
import sys
import openpyxl
import pathlib
import pprint

FORMAT = "[%(asctime)s %(filename)s->%(funcName)s():%(lineno)s]%(levelname)s: %(message)s"

logging.basicConfig(
    level=logging.INFO,
    filename="vane.log",
    filemode="w",
    format=FORMAT,
)


class XcelClient:
    """Creates an instance of the Excel Client."""

    def __init__(self, definitions_file):
        """Initializes the excel client

        Args:
            spreadsheet (obj): Spreadsheet
        """

        self.definitions = self._import_yaml(definitions_file)
        # pprint.pprint(self.definitions)
        self.spreadsheet = ""
        self.data_model = {"duts": []}

        xcel_defs = self.definitions["parameters"]["xcel_definitions"]
        self.xcel_defs = self._import_yaml(xcel_defs)

        xcel_schema = self.definitions["parameters"]["xcel_schema"]
        self.xcel_schema = self._import_yaml(xcel_schema)

    def _import_yaml(self, yaml_file):
        """Import YAML file as python data structure

        Args:
            yaml_file (str): Name of YAML file
        """

        logging.info(f"Opening {yaml_file} for read")
        try:
            with open(yaml_file, "r") as input_yaml:
                try:
                    yaml_data = yaml.safe_load(input_yaml)
                    logging.info(
                        f"Inputed the following yaml: {yaml_data}")
                    return yaml_data
                except yaml.YAMLError as err_data:
                    logging.error(f"Error in YAML file. {err_data}")
                    sys.exit(1)
        except OSError as err_data:
            logging.error(
                f"Defintions file: {yaml_file} not found. {err_data}"
            )
            sys.exit(1)

    def import_spreadsheet(self):
        """import spreadsheet for parsing

        return:
            lld_spreadsheet (obj):    Abstraction of spreadsheet,
        """

        spreadsheet_path = self.definitions["parameters"]["spreadsheet"]
        file_object = pathlib.Path(spreadsheet_path)

        if not file_object.exists():
            logging.error(f"Spreadsheet {spreadsheet_path} does not exist")
            print(f"SPREADSHEET {spreadsheet_path} DOES NOT EXIST")
            sys.exit(1)

        try:
            self.spreadsheet = openpyxl.load_workbook(spreadsheet_path)
            logging.info(
                f"Successfully importted spreadsheet {spreadsheet_path}"
            )
        except IOError:
            logging.error(f"Error opening spreadsheet {spreadsheet_path}")
            print(f"ERROR OPENING SPREADSHEET {spreadsheet_path}")
            sys.exit(1)

    def parse_spreadsheet(self):
        """Use excel definitioins to parse informations from spreadsheet"""

        logging.info(
            "Use excel definitioins to parse informations from spreadsheet"
        )
        worksheet = self._return_worksheet("HostnameMgmt")
        self._parse_host_tab(worksheet)
        pprint.pprint(self.data_model)

        worksheet = self._return_worksheet("wiremap")
        self._parse_wiremap_tab(worksheet)

    def _return_worksheet(self, ws_role):
        """Returns worksheet name

        Args:
            ws_role (str): Excel Worksheet role

        Returns:
            str: Excel worksheet name
        """

        tab = None
        worksheets = self.xcel_defs["worksheets"]
        ws_data = [x for x in worksheets if ws_role == x["role"]]
        logging.info(f"Found Worksheet role {ws_data}")
        ws_name = ws_data[0]["name"]
        logging.info(f"Worksheet name is {ws_name}")

        for worksheet in self.spreadsheet.sheetnames:
            logging.info(
                f"Comparing criteria |{ws_name}| to tab |{worksheet}|")
            if ws_name == worksheet:
                logging.info(f"Matched tab {worksheet}")
                tab = worksheet

        if tab:
            logging.info(f"Found tab {tab}")
            return tab
        else:
            logging.error(f"Error spreadsheet worksheet not found {ws_name}")
            print(f"ERROR NO SPREADSHEET TAB NOT FOUND {ws_name}")
            sys.exit(1)

    def _parse_host_tab(self, ws_name):
        """Parse host tabs in worksheet

        Args:
            ws_name (str): Name of excel worksheet
        """

        logging.info("Parse host tabs in worksheet")
        worksheets = self.xcel_defs["worksheets"]
        ws_data = [x for x in worksheets if ws_name == x["name"]]
        header_row = ws_data[0]["header_row"]

        host_col = header_row["Hostname"]["column"]
        end_row = len(self.spreadsheet[ws_name][host_col])
        end_col = len(header_row)
        worksheet = self.spreadsheet[ws_name]

        table_dimensions = {
            "start_row": 1,
            "end_row": end_row,
            "start_col": 1,
            "end_col": end_col,
            "header_row": header_row,
            "interval": 0,
            "seed_col": 0,
            "multiplier": 1,
            "device_start": 0,
        }

        table = self._read_spreadsheet(worksheet, table_dimensions)

        for row in table:
            if table[row]["hostname"]:
                self.data_model["duts"].append(table[row])

    def _return_cols(self, table_dimensions, multi_col):
        """Based on table dimensions return a range of cols

        Args:
            table_dimensions (dict): hash of table characteristics
            multi_col (int): table iteration

        Returns:
            (list): range of columns for iteration
        """

        cols = []
        start_col = table_dimensions["start_col"]
        seed_col = table_dimensions["seed_col"]
        interval = table_dimensions["interval"]

        cols += range(start_col, seed_col)
        logging.info(f"Add common columns {cols}")
        interval_start = seed_col * multi_col
        interval_end = (seed_col * multi_col) + interval
        cols += range(interval_start, interval_end)
        logging.info(f"Add interval columns {cols}")

        logging.info(f"Interesting columns {cols}")

        return cols

    def _read_spreadsheet(self, worksheet, table_dimensions):
        """Read spreadsheet and return data from it

        Args:
            worksheet (obj): spreadsheet worksheet to read
            table_dimensions (dict): hash of table characteristics

        Returns:
            (dict): table data
        """

        logging.info(f"Unpacking table dimensions {table_dimensions}")
        start_row = table_dimensions["start_row"]
        end_row = table_dimensions["end_row"]
        start_col = table_dimensions["start_col"]
        end_col = table_dimensions["end_col"]
        header_row = table_dimensions["header_row"]
        interval = table_dimensions["interval"]
        seed_col = table_dimensions["seed_col"]
        device_start = table_dimensions["device_start"]
        multiplier = table_dimensions["multiplier"] + 1

        table = {}
        counter = 1

        for multi_col in range(1, multiplier):
            logging.info(f"Table iteration: {multi_col}")
            cols = []

            if seed_col != 0:
                cols += self._return_cols(table_dimensions, multi_col)
            else:
                cols += range(start_col, end_col)

            logging.info(f"Interesting rows {range(start_row, end_row)}")

            for row in range(start_row, end_row):
                if row == start_row:
                    header_fields = []

                    for col in cols:
                        cell = worksheet.cell(row=row, column=col).value
                        header_fields.append(cell)

                    logging.info(f"Header Fields are {header_fields}")
                else:
                    table[counter] = {}

                    for col in cols:
                        if col < seed_col:
                            header_key = col - 1
                        else:
                            header_key = (col - 1) - (
                                (multi_col - 1) * interval
                            )

                        raw_key = header_fields[header_key]
                        key = header_row[raw_key]["role"]
                        cell = worksheet.cell(row=row, column=col).value

                        if header_key == (seed_col - 1) and device_start != 0:
                            device_col = seed_col * multi_col
                            device = worksheet.cell(
                                row=device_start, column=device_col
                            ).value
                            table[counter]["device"] = device

                        table[counter][key] = cell

                counter += 1

        logging.info(f"Raw table output: {table}")
        table = self._clean_table(table)

        return table

    def _parse_wiremap_tab(self, ws_name):
        """Parse wiremap worksheets

        Args:
            ws_name (str): Name of worksheet to pasrse
        """

        worksheets = self.xcel_defs["worksheets"]
        ws_data = [x for x in worksheets if ws_name == x["name"]]
        device_row = ws_data[0]["device_row"]
        device_start = device_row["start_row"]
        interval = device_row["interval"]
        multiplier = device_row["multiplier"]
        seed_col = device_row["column"]
        header_row = device_row["header_row"]

        int_col = header_row["Interface Number"]["column"]
        start_row = header_row["Interface Number"]["start_row"]
        end_row = len(self.spreadsheet[ws_name][int_col])
        end_col = len(header_row)
        worksheet = self.spreadsheet[ws_name]
        raw_start_col = header_row["Packet Processor"]["column"]
        start_col = ord(raw_start_col.lower()) - 96
        seed_col = ord(seed_col.lower()) - 96

        table_dimensions = {
            "start_row": start_row,
            "end_row": end_row,
            "start_col": start_col,
            "end_col": end_col,
            "header_row": header_row,
            "interval": interval,
            "seed_col": seed_col,
            "multiplier": multiplier,
            "device_start": device_start,
        }

        table = self._read_spreadsheet(worksheet, table_dimensions)

        pprint.pprint(table)

    def _clean_table(self, table):
        """Take table constructed from excel and remove rows with None values

        Args:
            table (dict): Raw excel translation to a dict

        Returns:
            dict: Table without None value rows
        """

        logging.info(f"Clean table {table}")
        new_table = {}

        for row in table:
            table_values = table[row]
            none_flag = True

            for key in table_values:
                value = table_values[key]

                if value is not None and key != "device":
                    none_flag = False

            if not none_flag:
                new_table[row] = table[row]

        logging.info(f"Cleaned table output: {table}")
        return new_table
