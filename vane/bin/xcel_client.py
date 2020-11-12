#!/usr/bin/python3
#
# Copyright (c) 2019, Arista Networks EOS+
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#
# * Redistributions of source code must retain the above copyright notice, this
#   list of conditions and the following disclaimer.
#
# * Redistributions in binary form must reproduce the above copyright notice,
#   this list of conditions and the following disclaimer in the documentation
#   and/or other materials provided with the distribution.
#
# * Neither the name of the Arista nor the names of its
#   contributors may be used to endorse or promote products derived from
#   this software without specific prior written permission.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
# FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
# DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
# SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
# CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
# OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
# OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
#

"""Utilities for using PyTest in network testing"""

import logging
import yaml
import sys
import openpyxl
import pathlib
import pprint


logging.basicConfig(level=logging.INFO, filename='vane.log', filemode='w',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


class XcelClient:
    """ Creates an instance of the Excel Client.
    """

    def __init__(self, definitions_file):
        """ Initializes the excel client

            Args:
                spreadsheet (obj): Spreadsheet 
        """

        self.definitions = self._import_yaml(definitions_file)
        self.spreadsheet = ""
        self.data_model = {"switches": []}

        xcel_defs = self.definitions['parameters']['xcel_definitions']
        self.xcel_defs = self._import_yaml(xcel_defs)

        xcel_schema = self.definitions['parameters']['xcel_schema']
        self.xcel_schema = self._import_yaml(xcel_schema)


    def _import_yaml(self, yaml_file):
        """ Import YAML file as python data structure

            Args:
                yaml_file (str): Name of YAML file
        """

        logging.info(f'Opening {yaml_file} for read')
        try:
            with open(yaml_file, 'r') as input_yaml:
                try:
                    yaml_data = yaml.safe_load(input_yaml)
                    logging.info(f'Inputed the following yaml: '
                                 f'{yaml_data}')
                    return yaml_data
                except yaml.YAMLError as err_data:
                    logging.error(f'Error in YAML file. {err_data}')
                    sys.exit(1)
        except OSError as err_data:
            logging.error(f'Defintions file: {yaml_file} not '
                          f'found. {err_data}')
            sys.exit(1)

    def import_spreadsheet(self):
        """ import spreadsheet for parsing

            return:
                lld_spreadsheet (obj):    Abstraction of spreadsheet,
        """

        spreadsheet_path = self.definitions['parameters']['spreadsheet']
        file_object = pathlib.Path(spreadsheet_path)

        if not file_object.exists():
            logging.error(f'Spreadsheet {spreadsheet_path} does not exist')
            print(f'SPREADSHEET {spreadsheet_path} DOES NOT EXIST')
            sys.exit(1)

        try:
            self.spreadsheet = openpyxl.load_workbook(spreadsheet_path)
            logging.info(f'Successfully importted spreadsheet {spreadsheet_path}')
        except:
            logging.error(f'Error opening spreadsheet {spreadsheet_path}')
            print(f'ERROR OPENING SPREADSHEET {spreadsheet_path}')
            sys.exit(1)

    def parse_spreadsheet(self):
        """ Use excel definitioins to parse informations from spreadsheet
        """

        worksheet = self._return_worksheet('HostnameMgmt')
        self._parse_host_tab(worksheet)
        worksheet = self._return_worksheet('wiremap')
        self._parse_wiremap_tab(worksheet)

    def _return_worksheet(self, ws_role):
        """ Returns worksheet name

        Args:
            ws_role (str): Excel Worksheet role

        Returns:
            str: Excel worksheet name
        """

        tab = None
        worksheets = self.xcel_defs['worksheets']
        ws_data = [x for x in worksheets if ws_role == x['role']]
        logging.info(f'Found Worksheet role {ws_data}')
        ws_name = ws_data[0]['name']
        logging.info(f'Worksheet name is {ws_name}')

        for worksheet in self.spreadsheet.sheetnames:
            logging.info(f'Comparing criteria |{ws_name}| to tab |{worksheet}|')
            if ws_name == worksheet:
                logging.info(f"Matched tab {worksheet}")
                tab = worksheet
        
        if tab:
            logging.info(f'Found tab {tab}')
            return tab
        else:
            logging.error(f'Error spreadsheet worksheet not found {ws_name}')
            print(f'ERROR NO SPREADSHEET TAB NOT FOUND {ws_name}')
            sys.exit(1)
                
    def _parse_host_tab(self, ws_name):

        worksheets = self.xcel_defs['worksheets']
        ws_data = [x for x in worksheets if ws_name == x['name']]
        header_row = ws_data[0]['header_row']

        print(header_row)

        host_col = header_row['Hostname']['column']
        end_row = len(self.spreadsheet[ws_name][host_col])
        end_col = len(header_row)
        worksheet = self.spreadsheet[ws_name]

        table_dimensions = {
            'start_row' : 1,
            'end_row': end_row,
            'start_col': 1,
            'end_col' : end_col,
            'header_row' : header_row,
            'interval' : 0,
            'seed_col' : 0,
            'multiplier' : 1,
            'device_start' : 0,
        }

        table = self._read_spreadsheet2(worksheet, table_dimensions)

        # print('>>> table')
        # pprint.pprint(table)

        for row in table:
            if table[row]['hostname']:
                self.data_model["switches"].append(table[row])

        #pprint.pprint(self.data_model)
        # hosts = [x for x in ws_data if 'hostname' == x['role']]
        # host_col = hosts['column']
        # hosts = [x for x in ws_data if 'hostname' == x['role']]
        # host_col = hosts['column']
# 
        # for header_field in header_row:
        #     name = header_field['name']
        #     column = header_field['column']
        #     last_row = len(self.spreadsheet[ws_name][column])
        #     print(f'{name} {column}')

    def _return_cols(self, table_dimensions, multi_col):

        cols = []
        start_col = table_dimensions['start_col']
        seed_col = table_dimensions['seed_col']
        interval = table_dimensions['interval']

        cols += range(start_col, seed_col)
        logging.info(f'Add common columns {cols}')
        interval_start = (seed_col * multi_col)
        interval_end = (seed_col * multi_col) + interval
        cols += range(interval_start, interval_end)
        logging.info(f'Add interval columns {cols}')

        logging.info(f'Interesting columns {cols}')

        return cols

    def _read_spreadsheet2(self, worksheet, table_dimensions):

        logging.info(f'Unpacking table dimensions {table_dimensions}')
        start_row = table_dimensions['start_row']
        end_row = table_dimensions['end_row']
        start_col = table_dimensions['start_col']
        end_col = table_dimensions['end_col']
        header_row = table_dimensions['header_row']
        interval = table_dimensions['interval']
        seed_col = table_dimensions['seed_col']
        device_start = table_dimensions['device_start']
        multiplier = (table_dimensions['multiplier'] + 1)

        table = {}
        counter = 1

        for multi_col in range(1, multiplier):
            logging.info(f'Table iteration: {multi_col}')
            cols = []

            if seed_col != 0:
                cols += self._return_cols(table_dimensions, multi_col)
            else:
                cols += range(start_col, end_col)   

            logging.info(f'Interesting rows {range(start_row, end_row)}')

            #for row in range(start_row, end_row):
            for row in range(start_row, end_row):
                # logging.info(f'Table iteration: {multi_col}, Table row iteration: {row}')

                if row == start_row:
                    header_fields = []
                    #for col in range(start_col, end_col):
                    for col in cols:
                        cell = worksheet.cell(row=row, column=col).value
                        # print(f'row: {row}, col: {col}, value: {cell}')
                        header_fields.append(cell)
                    logging.info(f'Header Fields are {header_fields}')
                else:
                    table[counter] = {}
                    #for col in range(start_col, end_col):
                    for col in cols:
                        #print(f'col - {col} / multi - {multiplier} = {col/multiplier}')
                        #col_key = int(col / multiplier)

                        if col < seed_col:
                            header_key = col - 1
                        else:
                            # header_key = ((col - 1) - (multi_col * interval))
                            # if interval == 0: interval = 1
                            header_key = ((col - 1) - ((multi_col - 1) * interval))

                            # logging.info(f'header_key {header_key} = col '
                            #              f'{col - 1} - (multi_col {multi_col}'
                            #              f' * interval {interval})')
                            #print(f'{header_key} | {col} | {multi_col} | {interval}')
                            # logging.info(f'\n>>>> header_key {header_key}\n'
                            #              f'>>>> col {col}\n'
                            #              f'>>>> multi_col {multi_col}\n'
                            #              f'>>>> interval {interval}\n'
                            #              f'>>>> {header_key} = (({col} - 1) - ({multi_col} * ({interval} - 1))')


                        # print(f'>>> {header_key}, {seed_col - 1}')

                        raw_key = header_fields[header_key]
                        key = header_row[raw_key]['role']
                        # logging.info(f'Header_key is {header_key}, Raw_Key is '
                        #              f'{raw_key}, Key is {key}')

                        cell = worksheet.cell(row=row, column=col).value
                        # logging.info(f'Table iteration: {multi_col}, Table '
                        #              f'row iteration: {row}, Table column '
                        #              f'iteration: {col}, Cell value: {cell}')
                        # logging.info(f'Adding {cell} to {key} in table using '
                        #              f'{header_key}')

                        if header_key == (seed_col - 1) and device_start != 0:
                            device_col = seed_col * multi_col
                            # print(f'>>> device_row: {device_start} device_col: {device_col}')
                            device = worksheet.cell(row=device_start, column=device_col).value
                            #print(f'>>> device: {device}')
                            table[counter]['device'] = device

                        table[counter][key] = cell
                        # logging.info(f'Table status: {table}')

                counter += 1

        logging.info(f'Raw table output: {table}')
        table = self._clean_table(table)

        # print('>>>> table')
        # pprint.pprint(table)

        return table

    def _parse_wiremap_tab(self, ws_name):

        worksheets = self.xcel_defs['worksheets']
        ws_data = [x for x in worksheets if ws_name == x['name']]
        device_row = ws_data[0]['device_row']
        device_start = ws_data[0]['device_row']['start_row']
        interval = ws_data[0]['device_row']['interval']
        multiplier = ws_data[0]['device_row']['multiplier']
        seed_col = ws_data[0]['device_row']['column']
        header_row = ws_data[0]['device_row']['header_row']
        # print('>>> header:')
        # pprint.pprint(header_row)

        int_col = header_row['Interface Number']['column']
        start_row = header_row['Interface Number']['start_row']
        end_row = len(self.spreadsheet[ws_name][int_col])
        end_col = len(header_row)
        worksheet = self.spreadsheet[ws_name]
        raw_start_col = header_row['Packet Processor']['column']
        start_col = (ord(raw_start_col.lower()) - 96)
        seed_col = (ord(seed_col.lower()) - 96)

        table_dimensions = {
            'start_row' : start_row, # first row
            'end_row': end_row,      # last row
            'start_col': start_col,  # first col
            'end_col' : end_col,     # last col
            'header_row' : header_row, 
            'interval' : interval,   
            'seed_col' : seed_col,
            'multiplier' : multiplier,
            'device_start' : device_start,
        }

        table = self._read_spreadsheet2(worksheet, table_dimensions)
        # print('>>> table')
        # pprint.pprint(table)

    def _clean_table(self, table):
        """ Take table constructed from excel and remove rows with None values

        Args:
            table (dict): Raw excel translation to a dict

        Returns:
            dict: Table without None value rows
        """

        logging.info(f'Clean table {table}')
        new_table = {}

        for row in table:
            table_values = table[row]
            none_flag = True

            for key in table_values:
                value = table_values[key]

                if value != None and key != "device":
                    none_flag = False
            
            if not none_flag:
                new_table[row] = table[row]

        logging.info(f'Cleaned table output: {table}')
        return new_table
