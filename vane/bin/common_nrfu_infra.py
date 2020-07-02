#!/usr/bin/python3
#
# Copyright (c) 2019, Arista Networks EOS+
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#
# * Redistributions of source code must retain the above copyright notice, this
#   list of conditions and the following disclaimer.
#
# * Redistributions in binary form must reproduce the above copyright notice,
#   this list of conditions and the following disclaimer in the documentation
#   and/or other materials provided with the distribution.
#
# * Neither the name of the Arista nor the names of its
#   contributors may be used to endorse or promote products derived from
#   this software without specific prior written permission.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
# FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
# DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
# SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
# CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
# OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
# OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
#

"""Utilities for using PyTest in network testing"""

from pprint import pprint
import concurrent.futures
import openpyxl
import jinja2
import pyeapi
import yaml
from . import definitions

#SPREADSHEET = "/cas/cas/spreadsheets/Carey.Test.xlsx"
JINJA2_TEMPLATE = "/cas/cas/jinja2_templates/eapi.conf.j2"
CONFIGLET_FILE = "/home/docker/.eapi.conf"
LOG_FILE = "/cas/cas/nrfu_logs/nrfu_test.log"


def import_spreadsheet():
    """ import spreadsheet for parsing

        return:
           xlsx_workbook (obj): Abstraction of spreadsheet,
    """

    try:
        xlsx_workbook = openpyxl.load_workbook(definitions.SPREADSHEET)
        return xlsx_workbook
    except BaseException as error:
        print(f">>>  Error spreadsheet failed to open with following error: \
{error}")
        raise


def parse_management_addressing(xlsx_workbook):
    """ parse worksheet HostnameMgmt IP & Router ID and model hostname and
        management IP address

        Args:
           xlsx_workbook (obj): Abstraction of spreadsheet

        Return:
           mgmt_addresses (list): model of management ip addresses

        Data Model:
        - hostname: <router name>
          mgmt_ip: <a.b.c.d>
    """
    start_row = 2
    mgmt_addresses = []

    for worksheet in xlsx_workbook.sheetnames:
        if 'HostnameMgmt' in worksheet:
            last_row = len(xlsx_workbook[worksheet]['D'])

            for row_index in range(start_row, last_row):
                hostname = xlsx_workbook[worksheet].cell(
                    row=row_index, column=4).value
                mgmt_ip = xlsx_workbook[worksheet].cell(
                    row=row_index, column=8).value

                if hostname:
                    mgmt_addresses_index = len(mgmt_addresses)
                    mgmt_addresses.append({})
                    mgmt_addresses[mgmt_addresses_index]['hostname'] \
                        = hostname
                    mgmt_addresses[mgmt_addresses_index]['mgmt_ip'] \
                        = mgmt_ip.split('/')[0]

    return mgmt_addresses


def return_connection_list(xlsx_workbook):
    """ parse worksheet HostnameMgmt IP & Router ID and model hostname and
        management IP address

        Args:
           xlsx_workbook (obj): Abstraction of spreadsheet

        Return:
           mgmt_addresses (list): model of management ip addresses

        Data Model:
        - hostname: <router name>
          mgmt_ip: <a.b.c.d>
    """
    start_row = 2
    connection_list = []

    for worksheet in xlsx_workbook.sheetnames:
        if 'HostnameMgmt' in worksheet:
            last_row = len(xlsx_workbook[worksheet]['D'])

            for row_index in range(start_row, last_row):
                hostname = xlsx_workbook[worksheet].cell(
                    row=row_index, column=4).value

                if hostname:
                    connection_list.append(hostname)

    return connection_list


def render_eapi_config(mgmt_addresses):
    """ Render .eapi.conf file so pytests can log into devices

        Args:
           mgmt_addresses (list): model of management ip addresses
    """

    try:
        with open(JINJA2_TEMPLATE, 'r') as jinja_file:
            jinja2_template = jinja_file.read()
            resource_file = jinja2.Environment().from_string(
                jinja2_template).render(
                    mgmt_addresses=mgmt_addresses,
                    username=definitions.USERNAME,
                    password=definitions.PASSWORD,
                    transport=definitions.TRANSPORT)
    except IOError:
        pass

    try:
        with open(CONFIGLET_FILE, 'w') as output_file:
            output_file.write(resource_file)
    except IOError:
        pass


def connect_to_duts(mgmt_addresses):
    """ Use eapi to connect to Arista switches for testing

        Args:
          mgmt_addresses (list): model of management ip addresses

        return:
          connections (list): List of dictionaries with connection and name
                              of DUTs
    """

    connections = []

    for mgmt_address in mgmt_addresses:
        #print(f">>>> |{mgmt_address['hostname']}| transport={definitions.TRANSPORT}, username={definitions.USERNAME}, password={definitions.PASSWORD}")
        connections_index = len(connections)
        connections.append({})
        connections[connections_index]['connection'] = \
            pyeapi.connect_to(mgmt_address['hostname'])
            #pyeapi.connect(host=mgmt_address['hostname'], transport=definitions.TRANSPORT, username=definitions.USERNAME, password=definitions.PASSWORD)

        connections[connections_index]['name'] = mgmt_address['hostname']

    return connections


def return_interfaces(dut, xlsx_workbook):
    """ parse spreadsheet for interface connections and return them to test

        Args:
            dut (str):           hostname of dut
            xlsx_workbook (obj): Abstraction of spreadsheet,

        return:
          interface_list (list): list of interesting interfaces based on
                                 PS LLD spreadsheet
    """

    start_row = 3
    interface_list = []

    for worksheet in xlsx_workbook.sheetnames:
        if 'Wiremap' in worksheet:
            last_row = len(xlsx_workbook[worksheet]['C'])

            for row_index in range(start_row, last_row):
                hostname = xlsx_workbook[worksheet].cell(
                    row=row_index, column=3).value
                media_type = xlsx_workbook[worksheet].cell(
                    row=row_index, column=4).value
                interface_name = xlsx_workbook[worksheet].cell(
                    row=row_index, column=5).value
                z_hostname = xlsx_workbook[worksheet].cell(
                    row=row_index, column=18).value
                z_interface_name = xlsx_workbook[worksheet].cell(
                    row=row_index, column=20).value
                interface_entry = {}

                if hostname == dut:
                    interface_entry['hostname'] = dut
                    interface_entry['interface_name'] = interface_name
                    interface_entry['z_hostname'] = z_hostname
                    interface_entry['z_interface_name'] = z_interface_name
                    interface_entry['media_type'] = media_type
                    interface_list.append(interface_entry)

    return interface_list


def open_log_file():
    """ Open log file for logging test show commands

        Args:
          LOG_FILE (str): path and name of log file
    """

    try:
        with open(LOG_FILE, 'w') as log_file:
            pass
    except BaseException as error:
        print(f">>>  Error opening log file: {error}")
        raise


def output_to_log_file(test_suite, test_name, hostname, output):
    """ Open log file for logging test show commands

        Args:
          LOG_FILE (str): path and name of log file
    """

    try:
        with open(LOG_FILE, 'a') as log_file:
            log_file.write("\n%s::%s[%s]:\n%s" % (
                test_suite, test_name, hostname, output))

    except BaseException as error:
        print(f">>>  Error opening log file: {error}")
        raise


def return_show_cmd_output(show_cmd, dut, test_suite, test_name):
    """ Return model data and text output from show commands and log text output.

        Args:
          show_cmd (str): show command
          dut (dict): Dictionary containing dut name and connection
          test_suite (str): test suite name
          test_name (str): test case name

        return:
          show_output (dict): json output of cli command
          show_output (dict): plain-text output of cli command
    """

    show_output = dut['connection'].enable(show_cmd)
    show_output_text = dut['connection'].run_commands(
        show_cmd, encoding='text')
    output_to_log_file(
        test_suite, test_name, dut['name'], show_output_text[0]['output'])

    return show_output, show_output_text


def return_dut_info(show_cmds, test_suite):
    """ Use PS LLD spreadsheet to find interesting duts and then execute
        inputted show commands on each dut.  Return structured data of
        dut's output data, hostname, and connection

        Args:
          show_cmds (str):  list of interesting show commands
          test_suite (str): test suite name

        return:
          duts (dict): structured data of dut's output data, hostname, and
                       connection
    """

    xlsx_workbook = import_spreadsheet()
    mgmt_addresses = parse_management_addressing(xlsx_workbook)
    render_eapi_config(mgmt_addresses)
    duts = connect_to_duts(mgmt_addresses)

    for dut in duts:
        dut["output"] = {}

        for show_cmd in show_cmds:
            function_def = f'test_{("_").join(show_cmd.split())}'
            dut["output"][show_cmd] = {}
            json_output, text_output = return_show_cmd_output(
                show_cmd, dut, test_suite, function_def)
            dut["output"][show_cmd]["json"] = json_output[0]["result"]
            dut["output"][show_cmd]["text"] = text_output[0]["output"]

    return duts


def return_dut_info_threaded(show_cmds, test_suite):
    """ Use PS LLD spreadsheet to find interesting duts and then execute
        inputted show commands on each dut.  Return structured data of
        dut's output data, hostname, and connection.  Using threading to
        make method more efficent.

        Args:
          show_cmds (str):  list of interesting show commands
          test_suite (str): test suite name

        return:
          duts (dict): structured data of duts output data, hostname, and
                       connection
    """

    xlsx_workbook = import_spreadsheet()
    mgmt_addresses = parse_management_addressing(xlsx_workbook)
    render_eapi_config(mgmt_addresses)
    duts = connect_to_duts(mgmt_addresses)
    workers = len(duts)

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        {executor.submit(
            dut_thread, dut, show_cmds, test_suite, xlsx_workbook): dut for dut in duts}

    return duts


def dut_thread(dut, show_cmds, test_suite, xlsx_workbook):
    """ Execute inputted show commands on dut.  Update dut structured data
        with show output.

        Args:

          dut (dict): structured data of a dut output data, hostname, and
          connection test_suite (str): test suite name
    """

    dut["output"] = {}

    for show_cmd in show_cmds:
        function_def = f'test_{("_").join(show_cmd.split())}'
        dut["output"][show_cmd] = {}
        dut["output"]["interface_list"] = return_interfaces(dut["name"], xlsx_workbook)
        json_output, text_output = return_show_cmd_output(
            show_cmd, dut, test_suite, function_def)
        dut["output"][show_cmd]["json"] = json_output[0]["result"]
        dut["output"][show_cmd]["text"] = text_output[0]["output"]

def import_test_definition():
    """ import spreadsheet for parsing

        return:
           test_definition (dict): Abstraction of testing parameters
    """

    try:
        with open(definitions.TEST_DEFINITION_FILE, 'r') as test_definition:
            return yaml.load(test_definition, Loader=yaml.FullLoader)
    except BaseException as error:
        print(f">>>  Error Test Definition failed to open with following error: \
{error}")
        raise

def generate_connection_list(test_definition):
    """ test_definition to create a connection_list for a list of ids
        that will identify individual test runs

        Args:
            test_definition (dict): Abstraction of testing parameters

        Returns:
            connection_list (list): Ordered list of DUT hostnames
    """

    return [dut['name'] for dut in test_definition['duts']]

def generate_dut_info_threaded(show_cmds, test_suite):
    """ Use PS LLD spreadsheet to find interesting duts and then execute
        inputted show commands on each dut.  Return structured data of
        dut's output data, hostname, and connection.  Using threading to
        make method more efficent.

        Args:
          show_cmds (str):  list of interesting show commands
          test_suite (str): test suite name

        return:
          duts (dict): structured data of duts output data, hostname, and
                       connection
    """

    test_definition = import_test_definition()
    #render_eapi_config(test_definition['duts'])
    duts = connect_to_duts(test_definition['duts'])
    workers = len(duts)

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        {executor.submit(
            dut_thread2, dut, show_cmds, test_suite, test_definition): dut for dut in duts}

    return duts

def dut_thread2(dut, show_cmds, test_suite, test_definition):
    """ Execute inputted show commands on dut.  Update dut structured data
        with show output.

        Args:
          dut (dict): structured data of a dut output data, hostname, and
          connection test_suite (str): test suite name
    """

    dut["output"] = {}

    for show_cmd in show_cmds:
        function_def = f'test_{("_").join(show_cmd.split())}'
        dut["output"][show_cmd] = {}
        dut["output"]["interface_list"] = generate_interface_list(dut["name"], test_definition)
        json_output, text_output = return_show_cmd_output(
            show_cmd, dut, test_suite, function_def)
        dut["output"][show_cmd]["json"] = json_output[0]["result"]
        dut["output"][show_cmd]["text"] = text_output[0]["output"]

def generate_interface_list(dut_name, test_definition):
    """ test_definition is used to createa a interface_list for active
        DUT interfaces and attributes

        Returns:
            interface_list (list): list of active DUT interfaces and attributes
    """

    dut_hostnames = [dut['name'] for dut in test_definition['duts']]
    dut_index = dut_hostnames.index(dut_name)
    interface_list = test_definition['duts'][dut_index]['test_criteria'][0]['criteria']
    
    return interface_list